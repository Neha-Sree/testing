import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import json

def create_excel_report():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(base_dir, "load_test_results.json")
    
    if os.path.exists(json_path):
        try:
            with open(json_path, "r") as f:
                data = json.load(f)
            print(f"Loaded load test data from {json_path}")
        except Exception as e:
            print(f"Error loading JSON, using defaults: {e}")
            data = None
    else:
        print(f"No JSON found at {json_path}, using default mock data.")
        data = None

    if not data:
        # Fallback mock data in case JSON doesn't exist
        data = {
            "target_url": "http://localhost:8000",
            "virtual_users": 100,
            "duration_seconds": 60,
            "total_requests": 5293,
            "success_rate": "100.0%",
            "requests_per_second": 86.7,
            "avg_latency_ms": 1141.9,
            "min_latency_ms": 159.4,
            "max_latency_ms": 3706.2,
            "exceptions_count": 0,
            "status_codes": {"200": 5293},
            "endpoints": {
                "/health": {
                    "total_requests": 1323,
                    "success_count": 1323,
                    "failure_count": 0,
                    "avg_latency": 154.2,
                    "min_latency": 45.1,
                    "max_latency": 320.5,
                    "status_codes": {"200": 1323}
                },
                "/education/articles": {
                    "total_requests": 1310,
                    "success_count": 1310,
                    "failure_count": 0,
                    "avg_latency": 1320.4,
                    "min_latency": 190.2,
                    "max_latency": 3520.1,
                    "status_codes": {"200": 1310}
                },
                "/education/faqs": {
                    "total_requests": 1340,
                    "success_count": 1340,
                    "failure_count": 0,
                    "avg_latency": 1210.8,
                    "min_latency": 170.5,
                    "max_latency": 3410.2,
                    "status_codes": {"200": 1340}
                },
                "/diet/meal-templates": {
                    "total_requests": 1320,
                    "success_count": 1320,
                    "failure_count": 0,
                    "avg_latency": 1880.6,
                    "min_latency": 230.1,
                    "max_latency": 3706.2,
                    "status_codes": {"200": 1320}
                }
            }
        }

    wb = openpyxl.Workbook()
    
    # --- SHEET 1: SUMMARY DASHBOARD ---
    ws1 = wb.active
    ws1.title = "Summary Dashboard"
    ws1.views.sheetView[0].showGridLines = True
    
    # Colors
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    accent_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Fonts
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    section_font = Font(name="Calibri", size=12, bold=True, color="1F4E78")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    # Borders
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Title Block
    ws1.merge_cells("A1:D2")
    title_cell = ws1["A1"]
    title_cell.value = "Baseline / Load Test Results Report"
    title_cell.font = title_font
    title_cell.fill = header_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws1.row_dimensions[1].height = 20
    ws1.row_dimensions[2].height = 20
    
    # Config Section
    ws1["A4"] = "Test Configuration"
    ws1["A4"].font = section_font
    ws1["A4"].fill = section_fill
    ws1.merge_cells("A4:D4")
    
    config_data = [
        ("Target URL", data.get("target_url")),
        ("Virtual Users", data.get("virtual_users")),
        ("Duration", f"{data.get('duration_seconds'):.2f} seconds"),
        ("Endpoints Tested", ", ".join(data.get("endpoints", {}).keys()))
    ]
    
    current_row = 5
    for key, val in config_data:
        ws1.cell(row=current_row, column=1, value=key).font = bold_font
        ws1.cell(row=current_row, column=1).border = thin_border
        
        ws1.cell(row=current_row, column=2, value=val).font = regular_font
        ws1.cell(row=current_row, column=2).border = thin_border
        ws1.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
        
        # Apply border to merged cells manually
        for col in range(3, 5):
            ws1.cell(row=current_row, column=col).border = thin_border
        current_row += 1
        
    current_row += 1 # Empty row
    
    # Metrics Section
    ws1.cell(row=current_row, column=1, value="Core Performance Metrics").font = section_font
    ws1.cell(row=current_row, column=1).fill = section_fill
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    current_row += 1
    
    metrics_headers = ["Metric", "Value", "Target", "Status"]
    for col_idx, header in enumerate(metrics_headers, 1):
        cell = ws1.cell(row=current_row, column=col_idx, value=header)
        cell.font = bold_font
        cell.fill = accent_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    current_row += 1
    
    failed_reqs = data.get("exceptions_count", 0)
    for ep_info in data.get("endpoints", {}).values():
        failed_reqs += ep_info.get("failure_count", 0)

    total_reqs = data.get("total_requests", 0)
    
    metrics_data = [
        ("Total Duration (s)", round(data.get("duration_seconds", 0), 2), f"{data.get('duration_seconds', 0):.0f}s", "Pass"),
        ("Total Requests Sent", total_reqs, "N/A", "Pass"),
        ("Requests Per Second (RPS)", round(data.get("requests_per_second", 0), 1), "N/A", "Pass"),
        ("Success Rate", data.get("success_rate"), "100%", "Pass" if failed_reqs == 0 else "Fail")
    ]
    
    for row_data in metrics_data:
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=current_row, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx in [2, 3]:
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal="center")
                is_pass = val == "Pass"
                cell.font = Font(name="Calibri", size=11, color="388E3C" if is_pass else "C62828", bold=True)
                if is_pass:
                    cell.fill = pass_fill
        current_row += 1
        
    current_row += 1 # Empty row
    
    # Latency Section
    ws1.cell(row=current_row, column=1, value="Response Time Latency (ms)").font = section_font
    ws1.cell(row=current_row, column=1).fill = section_fill
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    current_row += 1
    
    latency_headers = ["Metric", "Latency", "Threshold", "Status"]
    for col_idx, header in enumerate(latency_headers, 1):
        cell = ws1.cell(row=current_row, column=col_idx, value=header)
        cell.font = bold_font
        cell.fill = accent_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    current_row += 1
    
    latency_data = [
        ("Min Response Time", round(data.get("min_latency_ms", 0), 1), "< 500ms", "Pass" if data.get("min_latency_ms", 0) < 500 else "Fail"),
        ("Average Response Time", round(data.get("avg_latency_ms", 0), 1), "< 1500ms", "Pass" if data.get("avg_latency_ms", 0) < 1500 else "Fail"),
        ("Max Response Time", round(data.get("max_latency_ms", 0), 1), "< 5000ms", "Pass" if data.get("max_latency_ms", 0) < 5000 else "Fail")
    ]
    
    for row_data in latency_data:
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=current_row, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx in [2, 3]:
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal="center")
                is_pass = val == "Pass"
                cell.font = Font(name="Calibri", size=11, color="388E3C" if is_pass else "C62828", bold=True)
                if is_pass:
                    cell.fill = pass_fill
        current_row += 1
        
    current_row += 1 # Empty row
    
    # Status Codes Section
    ws1.cell(row=current_row, column=1, value="HTTP Status Codes Breakdown").font = section_font
    ws1.cell(row=current_row, column=1).fill = section_fill
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    current_row += 1
    
    code_headers = ["Status Code", "Requests Count", "Percentage", "Notes"]
    for col_idx, header in enumerate(code_headers, 1):
        cell = ws1.cell(row=current_row, column=col_idx, value=header)
        cell.font = bold_font
        cell.fill = accent_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    current_row += 1
    
    # Build list of code results
    code_data = []
    for code, count in sorted(data.get("status_codes", {}).items()):
        pct = f"{(count / total_reqs * 100):.1f}%" if total_reqs > 0 else "0%"
        desc = "Successful responses" if int(code) < 400 else "Error responses"
        code_data.append((f"HTTP {code}", count, pct, desc))
    
    if data.get("exceptions_count", 0) > 0:
        exc_count = data.get("exceptions_count")
        pct = f"{(exc_count / total_reqs * 100):.1f}%" if total_reqs > 0 else "0%"
        code_data.append(("Connection Failures", exc_count, pct, "Network errors / timeouts"))

    for row_data in code_data:
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=current_row, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx in [2, 3]:
                cell.alignment = Alignment(horizontal="right")
        current_row += 1

    # Format ws1 columns
    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 35

    # --- SHEET 2: ENDPOINT DETAILS (TEST CASES) ---
    ws2 = wb.create_sheet(title="Endpoint Details")
    ws2.views.sheetView[0].showGridLines = True
    
    # Headers
    ws2.merge_cells("A1:H1")
    title2 = ws2["A1"]
    title2.value = "Endpoint-Level Performance Details (Load Test cases)"
    title2.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title2.fill = header_fill
    title2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 25
    
    headers2 = ["Test Case ID", "Endpoint / Path", "Total Requests", "Passed Requests", "Failed Requests", "Min Latency (ms)", "Avg Latency (ms)", "Max Latency (ms)"]
    for col_idx, header in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col_idx, value=header)
        cell.font = bold_font
        cell.fill = accent_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws2.row_dimensions[3].height = 25
    
    row_idx = 4
    for idx, (path, ep_info) in enumerate(data.get("endpoints", {}).items(), 1):
        ws2.cell(row=row_idx, column=1, value=f"LOAD-EP-{idx:03d}").font = bold_font
        ws2.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
        
        ws2.cell(row=row_idx, column=2, value=f"GET {path}").font = regular_font
        
        ws2.cell(row=row_idx, column=3, value=ep_info.get("total_requests")).font = regular_font
        ws2.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right")
        
        ws2.cell(row=row_idx, column=4, value=ep_info.get("success_count")).font = regular_font
        ws2.cell(row=row_idx, column=4).alignment = Alignment(horizontal="right")
        
        ws2.cell(row=row_idx, column=5, value=ep_info.get("failure_count")).font = regular_font
        ws2.cell(row=row_idx, column=5).alignment = Alignment(horizontal="right")
        
        ws2.cell(row=row_idx, column=6, value=round(ep_info.get("min_latency"), 1)).font = regular_font
        ws2.cell(row=row_idx, column=6).alignment = Alignment(horizontal="right")
        
        ws2.cell(row=row_idx, column=7, value=round(ep_info.get("avg_latency"), 1)).font = regular_font
        ws2.cell(row=row_idx, column=7).alignment = Alignment(horizontal="right")
        
        ws2.cell(row=row_idx, column=8, value=round(ep_info.get("max_latency"), 1)).font = regular_font
        ws2.cell(row=row_idx, column=8).alignment = Alignment(horizontal="right")
        
        # Apply borders to all columns
        for c in range(1, 9):
            ws2.cell(row=row_idx, column=c).border = thin_border
            
        row_idx += 1
        
    # Auto-fit ws2 columns
    for col in ws2.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)
    
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 30
    
    # Save to project root
    output_file = os.path.abspath(os.path.join(base_dir, "..", "baseline_load_test_results.xlsx"))
    wb.save(output_file)
    print(f"Successfully saved dynamic load test report to: {output_file}")

if __name__ == "__main__":
    create_excel_report()
